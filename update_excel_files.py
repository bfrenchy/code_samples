"""
This module updates specified Excel workbooks with the latest figures from the database.
It maintains all existing formulas and other sheets in the workbooks.
The paths for the workbooks are defined in a central configuration file.
"""

# Import necessary libraries
import os
import pandas as pd
from openpyxl import load_workbook, Workbook

# Import required functions from other modules
from src.utils.data_export_funcs import basics_export, get_service_usage
from src.calculations.cost_calculator import main as calculate_cost
from src.calculations.usage_forecast import main as usage_forecast
from src.calculations.viewing_forecast import main as viewing_forecast

# Additional functions 
from src.utils.data_export_funcs import get_device_info

# Import configuration information
from src.utils.config import wb_update_paths, custom_colors

# Define columns for renaming
basic_columns = {
    # Original column names mapped to new column names...
}

cost_columns = {
    # Original column names mapped to new column names...
}

usage_columns = {
    # Original column names mapped to new column names...
}

viewing_columns = {
    # Original column names mapped to new column names...
}

device_columns = {
    # Original column names mapped to new column names...
}

service_usage_columns = {
    # Original column names mapped to new column names...
}

# Function to update or create sheets in each workbook
def update_sheets(df, sheet, file_path, wb):
    # Implementation of update_sheets function...

def main():
    basic_info = basics_export()
    basic_info = basic_info.rename(columns=basic_columns)  # user-friendly column names

    cost = calculate_cost()
    cost = cost.rename(columns=cost_columns)  # user-friendly column names

    usage = usage_forecast()
    usage = usage.rename(columns=usage_columns)  # user-friendly column names

    viewing = viewing_forecast()
    viewing = viewing.rename(columns=viewing_columns)  # user-friendly column names

    devices = get_device_info()
    devices = devices.rename(columns=device_columns)  # user-friendly column names

    service_usage = get_service_usage()
    service_usage = service_usage.rename(columns=service_usage_columns)  # user-friendly column names

    # Implementation for saving these dataframes to separate csv files...

    sheets = {
        'basic_info': [basic_info, 'Basic_Info'],
        'cost': [cost, 'Cost'],
        'usage': [usage, 'Usage_Forecast'],
        'viewing': [viewing, 'Viewing_Forecast'],
        'devices': [devices, 'Device_Info'],
        'service_usage': [service_usage, 'Service_Usage']
    }

    for path in wb_update_paths:
        print("")
        try:
            print(f"Updating workbook at {path}...")
            wb = load_workbook(path)
        except FileNotFoundError:
            print(f"Workbook not found at {path}. Skipping...")
            pass
        sheet_names = wb.sheetnames

        for sheet in sheets:
            if sheets[sheet][1] in sheet_names:
                update_sheets(sheets[sheet][0], sheets[sheet][1], path, wb)
            else:
                wb.create_sheet(sheets[sheet][1])
                update_sheets(sheets[sheet][0], sheets[sheet][1], path, wb)
            wb[sheets[sheet][1]].sheet_properties.tabColor = custom_colors[4][1:]
            wb.save(path)
    print("Update complete.")

# Test the module
main()
